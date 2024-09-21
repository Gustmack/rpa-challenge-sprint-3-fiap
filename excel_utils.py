import os
import openpyxl
from typing import List, Dict
# Função para salvar os dados no Excel


def salvar_dados_no_excel(nome_empresa: str, avaliacao: str, total_avaliacoes: str, salario: str, vagas: str,
                          entrevistas: str, beneficios: str, detalhes: Dict[str, List[str]]):
    # Obter o caminho absoluto da pasta atual do projeto
    caminho_arquivo = os.path.join(os.getcwd(), 'infojobs_empresas.xlsx')

    # Abrir o arquivo Excel
    if os.path.exists(caminho_arquivo):
        wb = openpyxl.load_workbook(caminho_arquivo)
    else:
        print("Arquivo Excel não encontrado.")
        return

    # Selecionar a aba InfoJobs para salvar as informações gerais
    if "InfoJobs" in wb.sheetnames:
        sheet_infojobs = wb["InfoJobs"]
    else:
        print("Aba 'InfoJobs' não encontrada.")
        return

    # Achar a próxima linha vazia na aba InfoJobs
    next_row_infojobs = 2  # Assumimos que o cabeçalho está na linha 1
    while sheet_infojobs.cell(row=next_row_infojobs, column=1).value is not None:
        next_row_infojobs += 1

    # Preencher os dados gerais na aba InfoJobs
    sheet_infojobs.cell(row=next_row_infojobs, column=1, value=nome_empresa)
    sheet_infojobs.cell(row=next_row_infojobs, column=2, value=avaliacao)
    sheet_infojobs.cell(row=next_row_infojobs, column=3, value=total_avaliacoes)
    sheet_infojobs.cell(row=next_row_infojobs, column=4, value=salario)
    sheet_infojobs.cell(row=next_row_infojobs, column=5, value=vagas)
    sheet_infojobs.cell(row=next_row_infojobs, column=6, value=entrevistas)
    sheet_infojobs.cell(row=next_row_infojobs, column=7, value=beneficios)

    # Verificar em qual aba salvar as avaliações detalhadas
    aba_destino = None
    if "nubank" in nome_empresa.lower():
        aba_destino = "Nubank"
    elif "c6 Bank" in nome_empresa.lower():
        aba_destino = "C6 Bank"
    elif "itau" in nome_empresa.lower() or "itaú" in nome_empresa.lower():
        aba_destino = "Itau"

    if aba_destino and aba_destino in wb.sheetnames:
        sheet_detalhes = wb[aba_destino]
    else:
        print(f"Aba '{aba_destino}' não encontrada.")
        return

    # Achar a próxima linha vazia na aba correspondente ao banco
    next_row_detalhes = 2
    while sheet_detalhes.cell(row=next_row_detalhes, column=1).value is not None:
        next_row_detalhes += 1

    # Preencher os dados das avaliações detalhadas
    for idx, (titulo, score, comentario, pros, contras, melhorias) in enumerate(
            zip(detalhes['avaliacao'], detalhes['score'], detalhes['comentarios'],
                detalhes['pros'], detalhes['contras'], detalhes['melhorias'])):
        sheet_detalhes.cell(row=next_row_detalhes + idx, column=1, value=nome_empresa)  # Nome da empresa na 1ª coluna
        sheet_detalhes.cell(row=next_row_detalhes + idx, column=2, value=titulo)        # Título da avaliação
        sheet_detalhes.cell(row=next_row_detalhes + idx, column=3, value=score)         # Score (estrelas)
        sheet_detalhes.cell(row=next_row_detalhes + idx, column=4, value=comentario)    # Comentário
        sheet_detalhes.cell(row=next_row_detalhes + idx, column=5, value=pros)          # Prós
        sheet_detalhes.cell(row=next_row_detalhes + idx, column=6, value=contras)       # Contras
        sheet_detalhes.cell(row=next_row_detalhes + idx, column=7, value=melhorias)     # Melhorias

    # Salvar o arquivo Excel
    wb.save(caminho_arquivo)
    print(f"Dados da empresa {nome_empresa} salvos com sucesso no Excel.")